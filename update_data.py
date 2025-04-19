import openpyxl
import re
import requests
from bs4 import BeautifulSoup
from datetime import datetime

def get_cftc_bitcoin_data():
    """Получает актуальные данные о позициях по Биткоину."""
    url = "https://www.cftc.gov/dea/futures/financial_lf.htm"
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    text = soup.get_text(separator='\n')

    # Регулярное выражение для поиска данных по биткойну
    bitcoin_pattern = re.compile(
        r'(BITCOIN - CHICAGO MERCANTILE EXCHANGE).*?'
        r'CFTC Code #\d+\s*Open Interest is\s*([\d,]+)\s*'
        r'Positions\s*([\d,]+)\s*([\d,]+)\s*([\d,]+)\s*'  # Dealer
        r'([\d,]+)\s*([\d,]+)\s*([\d,]+)\s*'  # Asset Manager
        r'([\d,]+)\s*([\d,]+)\s*([\d,]+)\s*'  # Leveraged Funds
        r'([\d,]+)\s*([\d,]+)\s*([\d,]+)\s*'  # Other Reportables
        r'([\d,]+)\s*([\d,]+)',  # Nonreportable
        re.DOTALL
    )

    match = bitcoin_pattern.search(text)
    
    if not match:
        print("Данные по Bitcoin на CME не найдены.")
        return None

    # Текущая дата
    html_text = response.text

    date_match = re.search(r"as of (\w+) (\d{1,2}), (\d{4})", html_text)
    if date_match:
        formatted_date = datetime.strptime(
            f"{date_match.group(1)} {date_match.group(2)}, {date_match.group(3)}", 
            "%B %d, %Y"
        ).strftime("%y%m%d")
        #print(formatted_date)
    else:
        print("Ошибка: дата не найдена.")
    current_date = datetime.now().strftime('%y%m%d')
    parsed_date = datetime.strptime(formatted_date, "%y%m%d")

# Изменяем год на 2025
    final_date = parsed_date.replace(year=2025)

    # Форматируем в нужный вид D.M.YYYY
    result = final_date.strftime("%d.%m.%Y")
    
    # Формируем данные в виде списка
    row_values = [
        "BITCOIN - CHICAGO MERCANTILE EXCHANGE",
        formatted_date,
        result,
        "133741",
        "CME",
        "00",
        "133",
        int(match.group(2).replace(',', '')),     # Open Interest
        int(match.group(3).replace(',', '')),     # Dealer Long All
        int(match.group(4).replace(',', '')),     # Dealer Short All
        int(match.group(5).replace(',', '')),     # Dealer Spreading All
        int(match.group(6).replace(',', '')),     # Asset Mgr Long All
        int(match.group(7).replace(',', '')),     # Asset Mgr Short All
        int(match.group(8).replace(',', '')),     # Asset Mgr Spreading All
        int(match.group(9).replace(',', '')),     # Lev Money Long All
        int(match.group(10).replace(',', '')),    # Lev Money Short All
        int(match.group(11).replace(',', '')),    # Lev Money Spreading All
        int(match.group(12).replace(',', '')),    # Other Rept Long All
        int(match.group(13).replace(',', '')),    # Other Rept Short All
        int(match.group(14).replace(',', '')),    # Other Rept Spreading All
        int(match.group(15).replace(',', '')),    # TotRept Long All
        int(match.group(16).replace(',', '')),    # TotRept Short All
        int(match.group(15).replace(',', '')),    # NonRept Long All
        int(match.group(16).replace(',', ''))     # NonRept Short All
    ]
    
    return row_values

def check_duplicate_date(ws, new_row):
    # Получаем все даты из столбца A (кроме заголовка)
    dates_column = [cell.value for cell in ws["B"]][1:]
    
    # Приводим даты к строковому формату (если они в datetime)
    dates_column_str = [
        date.strftime("%Y-%m-%d") if isinstance(date, datetime) else str(date) 
        for date in dates_column
    ]
    
    # Проверяем, есть ли новая дата в списке
    new_date = new_row[1]
    if isinstance(new_date, datetime):
        new_date_str = new_date.strftime("%Y-%m-%d")
    else:
        new_date_str = str(new_date)
    
    if new_date_str in dates_column_str:
        print(f"Данные за {new_date_str} уже существуют в файле.")
        return True
    else:
        return False

def update_history_file():
    """Обновляет файл history.xlsx, добавляя новые данные сверху."""
    filename = "history.xlsx"
    
    # Загружаем рабочую книгу и лист
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # Получаем свежую строку данных
    new_row = get_cftc_bitcoin_data()
    if not new_row:
        print("Не удалось получить новые данные.")
        return
    
    # Проверяем, не была ли эта дата уже занесена
    if check_duplicate_date(ws,new_row ):
        return

    # dates_column = [cell.value for cell in ws["A"]][1:]  # Исключаем заголовок
    # if new_row[1] in dates_column:
    #     print(f"Данные за {new_row[1]} уже существуют в файле.")
    #     return
    
    # Добавляем новую строку в начало листа (после заголовков)
    ws.append(new_row)
    
    # Сохраняем изменения
    wb.save(filename)
    print(f"Файл успешно обновлён. Новое значение за {new_row[1]}")

# Основной метод запуска обновления
#update_history_file()